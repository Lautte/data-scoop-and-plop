__author__ = 'laura'
'''  A Python script that uses the xlrd and openpyxl
frameworks to read user defined cells from a selected Excel spreadsheet and append
them to a seperate master workbook.
http://stackoverflow.com/questions/25207232/python-xlsx-writer-write-string-to-new-row
http://stackoverflow.com/questions/20736814/reading-a-cell-value-that-contains-a-formula-returns-0-0-when-using-xlrd
'''

import re #https://docs.python.org/2/howto/regex.html
#import shutil
#import os
import xlrd
from xlrd import open_workbook
from openpyxl import load_workbook
import openpyxl
import datetime


### Open 'Read' workbook with xlrd###
workbook = xlrd.open_workbook('4K3Q13.xls')

### Read data from source column using xlrd ###
def readsource(source, colnum):
    source = []
    for worksheet in workbook.sheets():
        for i in range(worksheet.nrows):
            if i>0:
                lstval = worksheet.cell_value(i, colnum)
                source += [lstval]
            #source = source[1:]
    return source
# Function for reading data from a selected column (colnum) to a predefined list (lst)
# Initialise empty lists which will store values
#Still trouble shooting this function as it keeps printing the first row for two of the three sheets
pond = []
t= []
flow = []
lab_pH = []
TSS = []
Fe=[]
Mn=[]
Se=[]
TSM=[]

#fill empty lists with respective data from readsource function
pond = readsource(pond, 2)
t = readsource(t, 3)
flow = readsource(flow, 4)
lab_pH = readsource(lab_pH, 5)
TSS= readsource(TSS, 6)
Fe= readsource(Fe, 7)
Mn= readsource(Mn, 8)
Se= readsource(Se, 9)
TSM=readsource(TSM,10)

#creates comprehensive list of tuples
d=zip(pond,t,flow,lab_pH,TSS,TSM,Fe,Mn,Se)

#creates a dictionary from a list of tuples using the first value in the tuple
#as the key (the pond name e.g. AI-12) and the other data in the tuples as values (i.e. date, flow, pH)
def create_dict(data):
    from collections import defaultdict
    ret = defaultdict(list)
    for v in data:
        ret[v[0]].append(v[1:])
    return ret

pond_dict=dict(create_dict(d)) #dictionary with the key being the pond's name

pond_dict.pop("", None) #this removes the key with "" from the pond dictionary

print "Cleaned up dict:",pond_dict





### Open destination workbook using openpyxl ###


destwb = load_workbook('ML_HD.xlsx')

'''for iterating through each sheet in the masterdocument to match the ponds in the
dictionary to the pond sheets in the masterdocument and then populate the rows in each respective
sheet with the surface water data'''
for SheetName in destwb.get_sheet_names():
    for key,value in pond_dict.iteritems():
        p=re.compile(key)
        m=p.search(SheetName)
        if m:
            print 'match found', m.group() #returns string matched by the RE
            sheet= destwb.get_sheet_by_name(SheetName)
            for lst in value:
                if type(lst[0])==unicode:
                    dt=lst[0]
                    #print "DICTIONARY DATE WAS IN UNICODE ",dt,type(dt)
                    dd=datetime.datetime.strptime(dt, '%m/%d/%y')
                    #print type(dd),dd
                    #datetime.toordinal()
# converting unicode to datetime.datetime format
                elif type(lst[0])==float:
                    #print "DATE IS IN FLOAT "
                    serial = lst[0]
                    seconds = (serial - 25569) * 86400.0
                    dd=datetime.datetime.utcfromtimestamp(seconds)
                    #dd = datetime.datetime.fromordinal(serial)
                    #print dd,type(dd)

                    ws=destwb.get_sheet_by_name(SheetName)
                    max_row=ws.get_highest_row()
                #starting from maximum row and going up searching for the same datetime.datetime
                #format
                for x in range(max_row,1,-1):
                    cll=sheet.cell(row=x, column=1)
                    c=cll.value
                    if type(c) is not type(dd):
                        continue
                        #print c," CELL type Does NOT Equal DICT type"
                        #print "type of dd=",type(dd)
                    elif type(c) is type(dd):
                        #print "CELL VALUE ",c,"<= DICT VALUE", dd
                        #print type(c), type(dd)," Time to do some WORK"
                        sheet.cell(row=x+1, column=1).value = dd #date
                        sheet.cell(row=x+1, column=5).value = lst[1]#flow
                        sheet.cell(row=x+1, column=6).value = lst[2]#lab pH
                        sheet.cell(row=x+1, column=9).value = lst[3]#TSS
                        sheet.cell(row=x+1, column=10).value = lst[4]#TSM
                        sheet.cell(row=x+1, column=11).value = lst[5]#Fe
                        sheet.cell(row=x+1, column=13).value = lst[6] #Mn
                        sheet.cell(row=x+1, column=15).value = lst[7] #Se
                        break
        elif not m:
            continue
            # print 'No Match'
            #print key,value
destwb.save('ML_HD_EDIT.xlsx')

""" This snippet of code is from here:  http://stackoverflow.com/questions/17299364/insert-row-into-excel-spreadsheet-using-openpyxl-in-python
Essentially, in order to "insert" a new row in excel you have to create a new copy of the file with the row inserted... :/

    old_sheet = destwb.get_sheet_by_name(SheetName)
    old_sheet.title = SheetName,'a'
    max_row = old_sheet.get_highest_row()
    max_col = old_sheet.get_highest_column()
    destwb.create_sheet(0, SheetName)
    new_sheet = wb.get_sheet_by_name(SheetName)

    # Do the header.
    for col_num in range(0, max_col):
        new_sheet.cell(row=0, column=col_num).value = old_sheet.cell(row=0, column=col_num).value

    # The row to be inserted. We're manually populating each cell.
    new_sheet.cell(row=1, column=0).value = 'DUMMY'
    new_sheet.cell(row=1, column=1).value = 'DUMMY'

    # Now do the rest of it. Note the row offset.
    for row_num in range(1, max_row):
        for col_num in range (0, max_col):
            new_sheet.cell(row = (row_num + 1), column = col_num).value = old_sheet.cell(row = row_num, column = col_num).value

    wb.save(file)  """
