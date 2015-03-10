__author__ = 'arthurl'
import re #https://docs.python.org/2/howto/regex.html
#import shutil
#import os
import xlrd
from xlrd import open_workbook
from openpyxl import load_workbook
import openpyxl
import datetime

### Open 'Read' workbook with xlrd###
workbook = xlrd.open_workbook('ML_HD.xls')
### Read data from source column using xlrd ###

### Open destination workbook using openpyxl ###
destwb = load_workbook('ML_HD_EDIT.xlsx')
wksht = destwb.create_sheet(0)

pointSheets = workbook.sheet_names()
r=0
for sheet in workbook.sheets():
    s = sheet.name
   # print s, type(s)
    source = []
    for rownum in range(8, sheet.nrows):
        r+=1
        source=[]
        s=str(s).split()
        s=s[0]
        source += [s]
        for colnum in range(0,16):
            lstval = sheet.cell_value(rownum, colnum)
            source += [lstval]

        if type(source[1]) is float:
            print "line:", source #here's where you need to insert it into an output file
            print "work it"

            for c in range(0,17):
                wksht.cell(row=r, column=c+1).value = source[c]
destwb.save('ML_HD_GIS.xlsx')

#print "Cleaned up dict:",pond_dict
